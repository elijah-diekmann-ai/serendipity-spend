from __future__ import annotations

import io
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import select

from serendipity_spend.core.db import SessionLocal
from serendipity_spend.core.storage import get_storage
from serendipity_spend.modules.claims.service import create_claim
from serendipity_spend.modules.documents.service import create_source_file
from serendipity_spend.modules.exports.models import ExportRun, ExportStatus
from serendipity_spend.modules.exports.service import create_export_run, generate_export
from serendipity_spend.modules.extraction.service import extract_source_file
from serendipity_spend.modules.identity.models import UserRole
from serendipity_spend.modules.identity.service import create_user


def _normalized_pdf_page_text(reader: PdfReader, page_idx: int) -> str:
    text = reader.pages[page_idx].extract_text() or ""
    return " ".join(text.split())


def _cell_style_signature(cell) -> tuple:
    fill = cell.fill
    fill_color = fill.fgColor
    if fill_color.type == "rgb":
        color_sig = ("rgb", fill_color.rgb)
    else:
        color_sig = (fill_color.type, fill_color.theme, fill_color.tint)
    border = cell.border
    align = cell.alignment
    font = cell.font
    return (
        cell.number_format,
        (font.name, font.bold, font.italic, font.size),
        (fill.patternType, color_sig),
        (border.left.style, border.right.style, border.top.style, border.bottom.style),
        (align.horizontal, align.vertical, align.wrap_text),
    )


def test_export_supporting_pdf_matches_reference_pdf_structure():
    with SessionLocal() as session:
        employee = create_user(
            session,
            email="dick.catlin@serendipitycapital.com",
            password="pw",
            role=UserRole.EMPLOYEE,
            full_name="Dick Catlin",
        )
        claim = create_claim(session, employee_id=employee.id, home_currency="SGD")
        claim_id = claim.id
        employee_id = employee.id

        pdf_bytes = Path("Data/DC__OOP__05 Sep 2025.pdf").read_bytes()
        source = create_source_file(
            session,
            claim=claim,
            user=employee,
            filename="DC__OOP__05 Sep 2025.pdf",
            content_type="application/pdf",
            body=pdf_bytes,
        )

        extract_source_file(source_file_id=str(source.id))

    run = create_export_run(claim_id=claim_id, requested_by_user_id=employee_id)
    generate_export(export_run_id=str(run.id))

    with SessionLocal() as session:
        run = session.scalar(select(ExportRun).where(ExportRun.id == run.id))
        assert run
        assert run.status == ExportStatus.COMPLETED
        assert run.summary_xlsx_key
        assert run.supporting_pdf_key
        assert run.supporting_zip_key is None

    supporting_bytes = get_storage().get(key=run.supporting_pdf_key)

    out_reader = PdfReader(io.BytesIO(supporting_bytes))
    ref_reader = PdfReader("Data/DC__OOP__05 Sep 2025.pdf")

    assert len(out_reader.pages) == len(ref_reader.pages)
    for idx in range(len(ref_reader.pages)):
        assert _normalized_pdf_page_text(out_reader, idx) == _normalized_pdf_page_text(
            ref_reader, idx
        )


def test_export_summary_xlsx_matches_reference_structure_and_styles():
    with SessionLocal() as session:
        employee = create_user(
            session,
            email="dick.catlin@serendipitycapital.com",
            password="pw",
            role=UserRole.EMPLOYEE,
            full_name="Dick Catlin",
        )
        claim = create_claim(session, employee_id=employee.id, home_currency="SGD")
        claim_id = claim.id
        employee_id = employee.id

        pdf_bytes = Path("Data/DC__OOP__05 Sep 2025.pdf").read_bytes()
        source = create_source_file(
            session,
            claim=claim,
            user=employee,
            filename="DC__OOP__05 Sep 2025.pdf",
            content_type="application/pdf",
            body=pdf_bytes,
        )
        extract_source_file(source_file_id=str(source.id))

    run = create_export_run(claim_id=claim_id, requested_by_user_id=employee_id)
    generate_export(export_run_id=str(run.id))

    with SessionLocal() as session:
        run = session.scalar(select(ExportRun).where(ExportRun.id == run.id))
        assert run
        assert run.status == ExportStatus.COMPLETED
        assert run.summary_xlsx_key

    summary_bytes = get_storage().get(key=run.summary_xlsx_key)
    out_wb = load_workbook(io.BytesIO(summary_bytes))
    out_ws = out_wb.active

    ref_wb = load_workbook("Data/Travel Reimbursement_DC_Jan 2026.xlsx")
    ref_ws = ref_wb.active

    assert sorted(str(r) for r in out_ws.merged_cells.ranges) == sorted(
        str(r) for r in ref_ws.merged_cells.ranges
    )

    for col in ("A", "B", "G", "H"):
        assert out_ws.column_dimensions[col].width == ref_ws.column_dimensions[col].width

    for addr in ("A1", "A2", "A3", "B1", "B2", "B3", "C5", "H5"):
        assert out_ws[addr].value == ref_ws[addr].value
        assert _cell_style_signature(out_ws[addr]) == _cell_style_signature(ref_ws[addr])

    # Column headers (row 6) and totals row formulas are part of the expected template structure.
    for col_idx in range(1, 9):
        addr = out_ws.cell(row=6, column=col_idx).coordinate
        assert out_ws[addr].value == ref_ws[addr].value
        assert _cell_style_signature(out_ws[addr]) == _cell_style_signature(ref_ws[addr])

    for addr in ("B27", "C27", "D27", "E27", "H27"):
        assert out_ws[addr].value == ref_ws[addr].value
        assert _cell_style_signature(out_ws[addr]) == _cell_style_signature(ref_ws[addr])


def test_export_supporting_pdf_includes_image_receipts():
    try:
        from PIL import Image
    except Exception:
        return

    with SessionLocal() as session:
        employee = create_user(
            session,
            email="employee@example.com",
            password="pw",
            role=UserRole.EMPLOYEE,
            full_name="Employee",
        )
        claim = create_claim(session, employee_id=employee.id, home_currency="SGD")
        claim_id = claim.id
        employee_id = employee.id

        img = Image.new("RGB", (120, 80), color=(255, 255, 255))
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        png_bytes = buf.getvalue()

        _ = create_source_file(
            session,
            claim=claim,
            user=employee,
            filename="receipt.png",
            content_type="image/png",
            body=png_bytes,
        )

    run = create_export_run(claim_id=claim_id, requested_by_user_id=employee_id)
    generate_export(export_run_id=str(run.id))

    with SessionLocal() as session:
        run = session.scalar(select(ExportRun).where(ExportRun.id == run.id))
        assert run
        assert run.status == ExportStatus.COMPLETED
        assert run.supporting_pdf_key

    supporting_bytes = get_storage().get(key=run.supporting_pdf_key)
    reader = PdfReader(io.BytesIO(supporting_bytes))
    assert len(reader.pages) == 1
