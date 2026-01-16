from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select

from serendipity_spend.core.db import SessionLocal
from serendipity_spend.modules.claims.models import ClaimStatus
from serendipity_spend.modules.claims.service import create_claim, submit_claim, update_claim
from serendipity_spend.modules.expenses.service import create_manual_item
from serendipity_spend.modules.identity.models import UserRole
from serendipity_spend.modules.identity.service import create_user
from serendipity_spend.modules.policy.models import (
    PolicyException,
    PolicyExceptionStatus,
    PolicyViolation,
    ViolationStatus,
)
from serendipity_spend.modules.policy.service import (
    decide_policy_exception,
    evaluate_claim,
    request_policy_exception,
)
from serendipity_spend.modules.workflow.models import Task, TaskStatus
from serendipity_spend.modules.workflow.service import approve_claim
from serendipity_spend.modules.workflow.models import ApprovalDecision


def test_submit_allows_exception_request_for_fail_rule():
    with SessionLocal() as session:
        employee = create_user(
            session, email="employee@example.com", password="pw", role=UserRole.EMPLOYEE
        )
        approver = create_user(
            session, email="approver@example.com", password="pw", role=UserRole.APPROVER
        )
        claim = create_claim(session, employee_id=employee.id, home_currency="USD")
        update_claim(
            session,
            claim=claim,
            user=employee,
            travel_start_date=date(2026, 1, 1),
            travel_end_date=date(2026, 1, 3),
            purpose="Work trip",
        )
        _ = create_manual_item(
            session,
            claim=claim,
            user=employee,
            vendor="Example Hotel",
            category="lodging",
            description="Hotel stay",
            transaction_date=date(2026, 1, 2),
            amount_original_amount=Decimal("650.00"),
            amount_original_currency="USD",
            metadata_json={"hotel_nights": 2, "employee_reviewed": True},
        )

        evaluate_claim(session, claim_id=claim.id)
        violation = session.scalar(
            select(PolicyViolation).where(
                PolicyViolation.claim_id == claim.id,
                PolicyViolation.rule_id == "R103",
                PolicyViolation.status == ViolationStatus.OPEN,
            )
        )
        assert violation

        with pytest.raises(HTTPException) as excinfo:
            submit_claim(session, claim=claim, user=employee)
        assert excinfo.value.status_code == 400
        assert "R103" in (excinfo.value.detail or {}).get("blocking_rules", [])

        exc = request_policy_exception(
            session,
            violation_id=violation.id,
            user=employee,
            justification="Conference hotel sold out; closest available option.",
        )
        assert exc.status == PolicyExceptionStatus.REQUESTED

        submitted = submit_claim(session, claim=claim, user=employee)
        assert submitted.status == ClaimStatus.NEEDS_APPROVER_REVIEW
        assert submitted.approver_id == approver.id

        # The corresponding policy task should be routed to the approver.
        task = session.scalar(
            select(Task).where(
                Task.claim_id == claim.id, Task.type == "POLICY_R103", Task.status == TaskStatus.OPEN
            )
        )
        assert task
        assert task.assigned_to_user_id == approver.id


def test_approver_must_decide_exception_before_approving_claim():
    with SessionLocal() as session:
        employee = create_user(
            session, email="employee@example.com", password="pw", role=UserRole.EMPLOYEE
        )
        approver = create_user(
            session, email="approver@example.com", password="pw", role=UserRole.APPROVER
        )
        claim = create_claim(session, employee_id=employee.id, home_currency="USD")
        update_claim(
            session,
            claim=claim,
            user=employee,
            travel_start_date=date(2026, 1, 1),
            travel_end_date=date(2026, 1, 3),
            purpose="Work trip",
        )
        _ = create_manual_item(
            session,
            claim=claim,
            user=employee,
            vendor="Example Hotel",
            category="lodging",
            description="Hotel stay",
            transaction_date=date(2026, 1, 2),
            amount_original_amount=Decimal("650.00"),
            amount_original_currency="USD",
            metadata_json={"hotel_nights": 2, "employee_reviewed": True},
        )

        evaluate_claim(session, claim_id=claim.id)
        violation = session.scalar(
            select(PolicyViolation).where(
                PolicyViolation.claim_id == claim.id,
                PolicyViolation.rule_id == "R103",
                PolicyViolation.status == ViolationStatus.OPEN,
            )
        )
        assert violation

        _ = request_policy_exception(
            session,
            violation_id=violation.id,
            user=employee,
            justification="Conference hotel sold out; closest available option.",
        )

        claim = submit_claim(session, claim=claim, user=employee)
        assert claim.status == ClaimStatus.NEEDS_APPROVER_REVIEW

        exc = session.scalar(
            select(PolicyException).where(
                PolicyException.claim_id == claim.id, PolicyException.status == PolicyExceptionStatus.REQUESTED
            )
        )
        assert exc

        with pytest.raises(HTTPException) as excinfo:
            approve_claim(
                session,
                claim=claim,
                user=approver,
                decision=ApprovalDecision.APPROVED,
                comment=None,
            )
        assert excinfo.value.status_code == 400

        exc = decide_policy_exception(
            session,
            exception_id=exc.id,
            user=approver,
            decision=PolicyExceptionStatus.APPROVED,
            comment="Approved with justification.",
        )
        assert exc.status == PolicyExceptionStatus.APPROVED

        claim = approve_claim(
            session,
            claim=claim,
            user=approver,
            decision=ApprovalDecision.APPROVED,
            comment=None,
        )
        assert claim.status == ClaimStatus.APPROVED


def test_export_xlsx_includes_policy_flags_column():
    from serendipity_spend.modules.exports.service import _build_reimbursement_xlsx

    with SessionLocal() as session:
        employee = create_user(
            session, email="employee@example.com", password="pw", role=UserRole.EMPLOYEE
        )
        claim = create_claim(session, employee_id=employee.id, home_currency="USD")
        update_claim(
            session,
            claim=claim,
            user=employee,
            travel_start_date=date(2026, 1, 1),
            travel_end_date=date(2026, 1, 3),
            purpose="Work trip",
        )
        item = create_manual_item(
            session,
            claim=claim,
            user=employee,
            vendor="Example Hotel",
            category="lodging",
            description="Hotel stay",
            transaction_date=date(2026, 1, 2),
            amount_original_amount=Decimal("650.00"),
            amount_original_currency="USD",
            metadata_json={"hotel_nights": 2, "employee_reviewed": True},
        )

        evaluate_claim(session, claim_id=claim.id)
        violation = session.scalar(
            select(PolicyViolation).where(
                PolicyViolation.claim_id == claim.id,
                PolicyViolation.rule_id == "R103",
                PolicyViolation.status == ViolationStatus.OPEN,
            )
        )
        assert violation
        _ = request_policy_exception(
            session,
            violation_id=violation.id,
            user=employee,
            justification="Conference hotel sold out; closest available option.",
        )

        evaluate_claim(session, claim_id=claim.id)
        open_violations = list(
            session.scalars(
                select(PolicyViolation).where(
                    PolicyViolation.claim_id == claim.id, PolicyViolation.status == ViolationStatus.OPEN
                )
            )
        )
        xlsx_bytes = _build_reimbursement_xlsx(
            claim=claim, employee=employee, items=[item], policy_violations=open_violations
        )

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.cell(row=6, column=9).value == "Policy flags"
    flags = ws.cell(row=8, column=9).value or ""
    assert "R103" in flags
    assert "EXC REQUESTED" in flags
