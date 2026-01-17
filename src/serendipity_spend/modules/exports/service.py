from __future__ import annotations

import io
import re
import time
import uuid
from copy import copy
from datetime import UTC, date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader, PdfWriter
from sqlalchemy import select

from serendipity_spend.core.db import SessionLocal
from serendipity_spend.core.logging import get_logger, log_event, log_exception, monotonic_ms
from serendipity_spend.core.storage import get_storage
from serendipity_spend.modules.claims.models import Claim
from serendipity_spend.modules.documents.models import EvidenceDocument, SourceFile
from serendipity_spend.modules.expenses.models import ExpenseItem, ExpenseItemEvidence
from serendipity_spend.modules.exports.models import ExportRun, ExportStatus
from serendipity_spend.modules.identity.models import User
from serendipity_spend.modules.policy.models import PolicyViolation, ViolationStatus
from serendipity_spend.modules.policy.service import evaluate_claim

logger = get_logger(__name__)


def create_export_run(*, claim_id: uuid.UUID, requested_by_user_id: uuid.UUID) -> ExportRun:
    with SessionLocal() as session:
        run = ExportRun(
            claim_id=claim_id,
            requested_by_user_id=requested_by_user_id,
            status=ExportStatus.QUEUED,
            error_message=None,
            summary_xlsx_key=None,
            supporting_pdf_key=None,
            supporting_zip_key=None,
            completed_at=None,
        )
        session.add(run)
        session.commit()
        session.refresh(run)
        log_event(
            logger,
            "export.run.created",
            claim_id=str(run.claim_id),
            export_run_id=str(run.id),
            requested_by_user_id=str(run.requested_by_user_id),
            status=run.status.value,
        )
        return run


def generate_export(*, export_run_id: str) -> None:
    with SessionLocal() as session:
        run = session.scalar(select(ExportRun).where(ExportRun.id == uuid.UUID(export_run_id)))
        if not run:
            return

        run.status = ExportStatus.RUNNING
        run.error_message = None
        session.add(run)
        session.commit()

        start = time.monotonic()
        log_event(
            logger,
            "export.generate.start",
            export_run_id=str(run.id),
            claim_id=str(run.claim_id),
        )
        try:
            claim = session.scalar(select(Claim).where(Claim.id == run.claim_id))
            if not claim:
                raise ValueError("Claim not found")

            employee = session.scalar(select(User).where(User.id == claim.employee_id))
            items = list(
                session.scalars(select(ExpenseItem).where(ExpenseItem.claim_id == claim.id))
            )
            sources = list(
                session.scalars(select(SourceFile).where(SourceFile.claim_id == claim.id))
            )

            evaluate_claim(session, claim_id=claim.id)
            policy_violations = list(
                session.scalars(
                    select(PolicyViolation).where(
                        PolicyViolation.claim_id == claim.id,
                        PolicyViolation.status == ViolationStatus.OPEN,
                    )
                )
            )
            xlsx_bytes = _build_reimbursement_xlsx(
                claim=claim, employee=employee, items=items, policy_violations=policy_violations
            )
            pdf_bytes = _build_supporting_pdf(session=session, claim=claim, sources=sources)

            summary_key = f"claims/{claim.id}/exports/{run.id}/summary.xlsx"
            supporting_key = f"claims/{claim.id}/exports/{run.id}/supporting.pdf"

            storage = get_storage()
            for key in (run.summary_xlsx_key, run.supporting_pdf_key):
                if not key:
                    continue
                try:
                    storage.delete(key=key)
                except Exception:
                    pass

            storage.put(key=summary_key, body=xlsx_bytes)
            storage.put(key=supporting_key, body=pdf_bytes)

            run.status = ExportStatus.COMPLETED
            run.summary_xlsx_key = summary_key
            run.supporting_pdf_key = supporting_key
            run.supporting_zip_key = None
            run.completed_at = datetime.now(UTC)
            session.add(run)
            session.commit()
            log_event(
                logger,
                "export.generate.finish",
                export_run_id=str(run.id),
                claim_id=str(run.claim_id),
                status=run.status.value,
                summary_key=run.summary_xlsx_key,
                supporting_key=run.supporting_pdf_key,
                duration_ms=monotonic_ms(start),
            )
        except Exception as e:  # noqa: BLE001
            run.status = ExportStatus.FAILED
            run.error_message = str(e)
            session.add(run)
            session.commit()
            log_exception(
                logger,
                "export.generate.error",
                export_run_id=str(run.id),
                claim_id=str(run.claim_id),
                duration_ms=monotonic_ms(start),
            )


def _build_reimbursement_xlsx(
    *,
    claim: Claim,
    employee: User | None,
    items: list[ExpenseItem],
    policy_violations: list[PolicyViolation] | None = None,
) -> bytes:
    wb = _load_reimbursement_template()
    ws = wb.active
    ws.title = "Reimbursement"
    employee_label = (
        (employee.full_name if employee and employee.full_name else None)
        or (employee.email if employee else None)
        or str(claim.employee_id)
    )
    ws["B1"] = employee_label
    if claim.travel_start_date and claim.travel_end_date:
        ws["B2"] = f"{claim.travel_start_date} to {claim.travel_end_date}"
    else:
        ws["B2"] = None
    ws["B3"] = claim.purpose or None

    ws["H6"] = claim.home_currency.upper()

    currency_col = {"USD": 3, "SGD": 4, "CAD": 5, "GBP": 6}
    policy_col = 9
    ws.cell(row=6, column=policy_col, value="Policy flags")
    ws.column_dimensions["I"].width = 45

    violations_by_item_id: dict[uuid.UUID, list[PolicyViolation]] = {}
    for v in policy_violations or []:
        if v.expense_item_id is None:
            continue
        violations_by_item_id.setdefault(v.expense_item_id, []).append(v)

    data_start_row = 8
    total_row = _find_total_row(ws) or (data_start_row + max(len(items), 1))
    max_data_rows = max(total_row - data_start_row, 0)

    if len(items) > max_data_rows and max_data_rows > 0:
        extra_rows = len(items) - max_data_rows
        _insert_styled_rows(
            ws,
            insert_at_row=total_row,
            count=extra_rows,
            style_from_row=total_row - 1,
        )
        total_row += extra_rows
        _update_total_formulas(ws, total_row=total_row, data_start_row=data_start_row)

    row = data_start_row
    for item in sorted(items, key=lambda i: (i.transaction_date or date.min, i.created_at)):
        ws.cell(row=row, column=1, value=item.transaction_date)
        ws.cell(row=row, column=2, value=item.description or item.vendor)

        col = currency_col.get(item.amount_original_currency.upper())
        if col:
            ws.cell(row=row, column=col, value=float(item.amount_original_amount))

        if item.fx_rate_to_home is not None:
            ws.cell(row=row, column=7, value=float(item.fx_rate_to_home))

        if item.amount_home_amount is not None:
            ws.cell(row=row, column=8, value=float(item.amount_home_amount))

        flags = violations_by_item_id.get(item.id) or []
        if flags:
            ws.cell(row=row, column=policy_col, value=_format_policy_flags(flags))

        row += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _load_reimbursement_template() -> Workbook:
    template_path = (
        Path(__file__).resolve().parents[4] / "Data" / "Travel Reimbursement_DC_Jan 2026.xlsx"
    )
    if template_path.exists():
        return load_workbook(template_path)

    # Fall back to a barebones workbook if the template is unavailable.
    wb = Workbook()
    wb.active.title = "Reimbursement"
    wb.active["A1"] = "Employee"
    wb.active["A2"] = "Travel period"
    wb.active["A3"] = "Purpose of the trip"
    headers = [
        "Date",
        "Description",
        "USD",
        "SGD",
        "CAD",
        "GBP",
        "EX rate",
        "Reimbursement",
        "Policy flags",
    ]
    for col, value in enumerate(headers, start=1):
        wb.active.cell(row=6, column=col, value=value)
    return wb


def _find_total_row(ws) -> int | None:
    # The provided template places a "Total" label in column B.
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == "Total":
            return row
    return None


def _insert_styled_rows(*, ws, insert_at_row: int, count: int, style_from_row: int) -> None:
    if count <= 0:
        return
    ws.insert_rows(insert_at_row, amount=count)
    for offset in range(count):
        target_row = insert_at_row + offset
        _copy_row_style(ws, source_row=style_from_row, target_row=target_row)


def _copy_row_style(*, ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        target_cell._value = None  # noqa: SLF001
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def _update_total_formulas(*, ws, total_row: int, data_start_row: int) -> None:
    data_end_row = max(total_row - 1, data_start_row)
    for col_letter in ("C", "D", "E", "F", "H"):
        cell = ws[f"{col_letter}{total_row}"]
        if isinstance(cell.value, str) and cell.value.startswith("=SUM("):
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"


def _format_policy_flags(violations: list[PolicyViolation]) -> str:
    severity_order = {"FAIL": 0, "NEEDS_INFO": 1, "WARN": 2, "PASS": 3}

    def one(v: PolicyViolation) -> str:
        parts = [v.rule_id, v.severity.value]
        exc = (v.data_json or {}).get("exception") or {}
        exc_status = str(exc.get("status") or "").strip()
        if exc_status:
            parts.append(f"EXC {exc_status}")
        elif v.severity.value != "FAIL" and bool((v.data_json or {}).get("submit_blocking")):
            parts.append("BLOCKS")
        return f"{' '.join(parts)} - {v.title}"

    return "; ".join(
        one(v)
        for v in sorted(
            violations, key=lambda x: (severity_order.get(x.severity.value, 99), x.rule_id)
        )
    )


def _build_supporting_pdf(
    *, session, claim: Claim, sources: list[SourceFile]
) -> bytes:
    writer = PdfWriter()
    sorted_sources = _sort_sources_for_supporting_pdf(
        session=session, claim=claim, sources=sources
    )
    for source in sorted_sources:
        body = get_storage().get(key=source.storage_key)
        reader = _source_file_to_pdf_reader(source=source, body=body)
        writer.append_pages_from_reader(reader)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _sort_sources_for_supporting_pdf(
    *, session, claim: Claim, sources: list[SourceFile]
) -> list[SourceFile]:
    source_ids = [s.id for s in sources]
    if not source_ids:
        return []

    # Prefer chronological ordering based on extracted items (if available), otherwise fall back
    # to upload time.
    source_to_min_date: dict[uuid.UUID, date] = {}
    rows = session.execute(
        select(EvidenceDocument.source_file_id, ExpenseItem.transaction_date)
        .join(
            ExpenseItemEvidence,
            ExpenseItemEvidence.evidence_document_id == EvidenceDocument.id,
        )
        .join(ExpenseItem, ExpenseItem.id == ExpenseItemEvidence.expense_item_id)
        .where(ExpenseItem.claim_id == claim.id, EvidenceDocument.source_file_id.in_(source_ids))
    ).all()
    for source_file_id, tx_date in rows:
        if tx_date is None:
            continue
        existing = source_to_min_date.get(source_file_id)
        if existing is None or tx_date < existing:
            source_to_min_date[source_file_id] = tx_date

    def sort_key(s: SourceFile) -> tuple[date, datetime, uuid.UUID]:
        tx_date = source_to_min_date.get(s.id, date.max)
        return (tx_date, s.created_at, s.id)

    return sorted(sources, key=sort_key)


def _source_file_to_pdf_reader(*, source: SourceFile, body: bytes) -> PdfReader:
    from serendipity_spend.modules.extraction.service import _detect_file_kind

    kind = _detect_file_kind(
        filename=source.filename,
        content_type=source.content_type,
        body=body,
    )
    if kind == "bad_pdf_upload":
        raise ValueError(
            "Bad upload: file looks like a PDF but does not start with the %PDF header "
            f"({source.filename})"
        )
    if kind == "pdf":
        return PdfReader(io.BytesIO(body))
    if kind == "image":
        pdf_bytes = _image_bytes_to_pdf(body)
        return PdfReader(io.BytesIO(pdf_bytes))
    if kind == "text":
        pdf_bytes = _text_bytes_to_pdf(
            body,
            filename=source.filename,
            content_type=source.content_type,
        )
        return PdfReader(io.BytesIO(pdf_bytes))
    raise ValueError(f"Unsupported supporting document type: {source.filename}")


def _text_bytes_to_pdf(body: bytes, *, filename: str, content_type: str | None) -> bytes:
    ctype = (content_type or "").lower()
    is_html = ctype.startswith("text/html") or filename.lower().endswith((".html", ".htm"))
    try:
        text = body.decode("utf-8", errors="replace")
    except Exception:
        text = body.decode("latin-1", errors="replace")
    if is_html:
        text = _html_to_text(text)
    return _text_to_pdf(text)


def _html_to_text(html: str) -> str:
    from html import unescape

    html = re.sub(r"(?is)<(script|style).*?>.*?</\1>", "", html)
    html = re.sub(r"(?i)<br\s*/?>", "\n", html)
    html = re.sub(r"(?i)</p\s*>", "\n\n", html)
    html = re.sub(r"(?i)</div\s*>", "\n", html)
    html = re.sub(r"(?s)<[^>]+>", "", html)
    html = unescape(html)
    lines = [re.sub(r"\s+", " ", ln).strip() for ln in html.splitlines()]
    out_lines: list[str] = []
    last_blank = False
    for ln in lines:
        if not ln:
            if not last_blank:
                out_lines.append("")
            last_blank = True
            continue
        out_lines.append(ln)
        last_blank = False
    return "\n".join(out_lines).strip()


def _text_to_pdf(text: str) -> bytes:
    # Minimal, dependency-free text->PDF using a built-in Type1 font.
    page_width = 612
    page_height = 792
    margin_x = 72
    margin_top = 72
    margin_bottom = 72
    font_size = 10
    leading = 12

    usable_width = page_width - 2 * margin_x
    char_width = font_size * 0.6  # Courier is 600 units wide.
    max_chars = max(int(usable_width / char_width), 40)

    def wrap_line(line: str) -> list[str]:
        line = line.replace("\t", "    ").rstrip("\n")
        if not line:
            return [""]
        out: list[str] = []
        while len(line) > max_chars:
            out.append(line[:max_chars])
            line = line[max_chars:]
        out.append(line)
        return out

    wrapped: list[str] = []
    for ln in (text or "").splitlines():
        wrapped.extend(wrap_line(ln))
    if not wrapped:
        wrapped = [""]

    lines_per_page = max(int((page_height - margin_top - margin_bottom) / leading), 1)
    pages = [
        wrapped[i : i + lines_per_page] for i in range(0, len(wrapped), lines_per_page)
    ]

    def pdf_escape(s: str) -> str:
        s = s.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        return "".join(ch if ord(ch) >= 32 else " " for ch in s)

    objects: list[bytes] = []
    # 1: catalog, 2: pages, 3: font
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    objects.append(b"")  # placeholder for /Pages
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>")

    kids: list[str] = []
    for idx, page_lines in enumerate(pages):
        page_obj_num = 4 + idx * 2
        content_obj_num = 5 + idx * 2
        kids.append(f"{page_obj_num} 0 R")

        start_y = page_height - margin_top - font_size
        stream_lines = ["BT", f"/F1 {font_size} Tf", f"{margin_x} {start_y} Td"]
        for ln in page_lines:
            stream_lines.append(f"({pdf_escape(ln)}) Tj")
            stream_lines.append(f"0 -{leading} Td")
        stream_lines.append("ET")
        stream = ("\n".join(stream_lines) + "\n").encode("latin-1", errors="replace")

        page_obj = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] "
            f"/Resources << /Font << /F1 3 0 R >> >> /Contents {content_obj_num} 0 R >>"
        ).encode("ascii")
        content_obj = (
            b"<< /Length "
            + str(len(stream)).encode("ascii")
            + b" >>\nstream\n"
            + stream
            + b"endstream"
        )
        objects.append(page_obj)
        objects.append(content_obj)

    pages_obj = (
        f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(pages)} >>"
    ).encode("ascii")
    objects[1] = pages_obj

    out = bytearray()
    out.extend(b"%PDF-1.4\n")

    offsets: list[int] = [0]
    for i, obj in enumerate(objects, start=1):
        offsets.append(len(out))
        out.extend(f"{i} 0 obj\n".encode("ascii"))
        out.extend(obj)
        out.extend(b"\nendobj\n")

    xref_offset = len(out)
    out.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    out.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        out.extend(f"{off:010d} 00000 n \n".encode("ascii"))
    out.extend(b"trailer\n")
    out.extend(f"<< /Size {len(objects) + 1} /Root 1 0 R >>\n".encode("ascii"))
    out.extend(b"startxref\n")
    out.extend(f"{xref_offset}\n".encode("ascii"))
    out.extend(b"%%EOF\n")
    return bytes(out)


def _image_bytes_to_pdf(body: bytes) -> bytes:
    try:
        from PIL import Image, ImageOps
    except Exception as e:  # noqa: BLE001
        raise ValueError("Pillow is required to convert images to PDF") from e

    img = Image.open(io.BytesIO(body))

    frames: list[Image.Image] = []
    n_frames = getattr(img, "n_frames", 1)
    for idx in range(n_frames):
        try:
            img.seek(idx)
        except Exception:
            break
        frame = img.copy()
        frame = ImageOps.exif_transpose(frame)
        if frame.mode not in {"RGB", "L"}:
            frame = frame.convert("RGB")
        frames.append(frame)

    if not frames:
        raise ValueError("Image could not be decoded")

    first, rest = frames[0], frames[1:]
    out = io.BytesIO()
    if rest:
        first.save(out, format="PDF", save_all=True, append_images=rest)
    else:
        first.save(out, format="PDF")
    return out.getvalue()
